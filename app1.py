import openpyxl


book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Cursos')
Curses_page = book['Cursos']

Curses_page.append(['CURSO DE AURICULOTERAPIA', '03/11'])
Curses_page.append(['CURSO DE HIDROLIPOCLASIA NÃO ASPIRATIIVA', '03/11'])
Curses_page.append(['CURSO DE MICROAGULHAMENTO: FACIAL E CORPORAL', '16/11'])
Curses_page.append(['CURSO DE LIBERAÇÃO MIOFASCIAL: MANUAL E INSTRUMENTAL', '16/11'])
Curses_page.append(['CURSO DE DRENAGEM LINFÁTICA INJETADA', '16/11'])
Curses_page.append(['CURSO DE LIMPEZA DE PELE + DERMAPLANING E HIDRAGLOSS', '17/11'])
Curses_page.append(['CURSO DE TOXINA BOTULINICA', '17/11'])
Curses_page.append(['CURSO DE BANDAGEM NEUROMUSCULAR', '17/11'])
Curses_page.append(['CURSO DE CINESIOTERAPIA E BIOMECÂNICA', '23/11'])
Curses_page.append(['CURSO DE DRY NEELING ASSOCIADO A ELETROTERAPIA', '23/11'])
Curses_page.append(['CURSO DE LIPOENZIMÁTICA DE PAPADA + SKINBOOSTER FACIAL + INTRADERMOTERAPIA CAPILAR', '23/11'])
Curses_page.append(['CURSO DE DRENAGEM LINFATICA + MASSAGEM MODELADORA + ULTRASSOM ESTETICO', '24/11'])
Curses_page.append(['CURSO DE PILATES APLICADO A PESSOA IDOSA  ', '24/11'])
Curses_page.append(['CURSO DE LASER TAPING NA GESTAÇÃO E NO PÓS-OPERATÓRIO', '24/11'])
Curses_page.append(['CURSO DE AVALIAÇÃO FISIOTERAPÊUTICA', '30/11'])
Curses_page.append(['CURSO DE PROCEDIMENTO ESTETICO INJETAVEL PARA MICROVASOS - PEIM', '30/11'])
Curses_page.append(['CURSO DE JATO DE PLASMA', '30/11'])

book.save("Cursoss.xlsx")